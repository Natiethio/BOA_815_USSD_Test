import unittest
import time
import pandas as pd
from bs4 import BeautifulSoup
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from appium.options.android import UiAutomator2Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import re
import sys
import os
from ATM_Withdrawal.atm_withdrawal import atm_withdrawal
from USSD_815__Test.USSD_815__Test2.USSD_815__Test.Airtime.airtimeupd import airtime_topup_for_ethio_telecom, airtime_topup_for_safaricom
from Transfer_to_own.transfer_to_own import transfer_to_own_account
import subprocess
import android_helper
def get_connected_device():
    try:
        result = subprocess.run(['adb', 'devices'], capture_output=True, text=True, check=True)
        output_lines = result.stdout.strip().split('\n')
        if len(output_lines) > 1:
            device_line = output_lines[1].split('\t')
            if len(device_line) == 2 and device_line[1] == 'device':
                return device_line[0]
        return None
    except subprocess.CalledProcessError as e:
        print(f"Error running ADB devices: {e}")
        return None

def get_android_version(device_id):
    try:
        result = subprocess.run(['adb', '-s', device_id, 'shell', 'getprop', 'ro.build.version.release'],
                              capture_output=True, text=True, check=True)
        return result.stdout.strip()
    except subprocess.CalledProcessError as e:
        print(f"Error getting Android version for {device_id}: {e}")
        return None

class TestAppium(unittest.TestCase):
    # Get device info automatically
    auto_device_id = android_helper.get_connected_device()
    auto_android_ver = android_helper.get_android_version(auto_device_id) if auto_device_id else None
    pkg, main_activity = android_helper.get_dialer_info()
    # Use command line args if provided, otherwise use auto-detected values, otherwise use defaults
    # device_id      = sys.argv[1] if len(sys.argv) > 1 else (auto_device_id if auto_device_id else "qewrtwtey")
    # android_ver    = sys.argv[2] if len(sys.argv) > 2 else (auto_android_ver if auto_android_ver else "12")
    pin            = sys.argv[1] if len(sys.argv) > 1 else "2073"
    account_number = sys.argv[2] if len(sys.argv) > 2 else "174615624"
    amount         = sys.argv[3] if len(sys.argv) > 3 else "10"
    phone_number   = sys.argv[4] if len(sys.argv) > 4 else "0970951608"
    safaricom_number = sys.argv[5] if len(sys.argv) > 5 else "0712911008"

    print(f"Using Device ID: {auto_device_id if auto_device_id else "qewrtwtey"}")
    print(f"Using Android Version: {auto_android_ver if auto_android_ver else "12"}")
    print(f"Using package: {pkg}")
    print(f"Using Main_Activity: {main_activity}")

    # [Rest of your existing TestAppium class code remains the same]

    def setUp(self) -> None:
        options = UiAutomator2Options()
        options.platform_name = "Android"
        options.automation_name = "UiAutomator2"
        # options.device_name = "R9ZR601C18H"
        # options.udid = "R9ZR601C18H"
        options.device_name = self.auto_device_id 
        options.udid = self.auto_device_id
        options.platform_version = self.auto_android_ver
        options.app_package = self.pkg
        options.app_activity = self.main_activity
        options.no_reset = True
        options.language = "en"
        options.locale = "US"
        
        try:
            self.driver = webdriver.Remote("http://127.0.0.1:4723", options=options)
            self.driver.implicitly_wait(5)
        except Exception as e:
           print("Error initializing appium",e)

    def tearDown(self) -> None:
        if self.driver:
            self.driver.quit()

    def dialussdcode(self,expected_texts, retries=5, wait_time=7):
       
        ussd_code = '*815' + '%23'

        for attempt in range(1, retries + 1):
            print(f"Dialing USSD code (Attempt {attempt})...", flush=True)
            self.driver.execute_script('mobile: shell', {
                'command': 'am',
                'args': ['start', '-a', 'android.intent.action.CALL', 'tel:' + ussd_code]
            })

            try:
                message_element = WebDriverWait(self.driver, wait_time).until(
                    EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
                )
                ussd_text = message_element.text.strip()
                print("USSD Response:", ussd_text, flush=True)

                if all(expected.lower() in ussd_text.lower() for expected in expected_texts):
                    print("Expected USSD page matched.", flush=True)
                    return True
                else:
                    self.cancel_ussd()
                    time.sleep(0.3)

                print("Expected USSD response not found. Retrying...", flush=True)

            except Exception as e:
                print(f"Error waiting for USSD response: {e}", flush=True)

            time.sleep(2)  

        self.fail(" Failed to load the expected USSD screen after retries.")


    def enter_pin_to_login(self):
        # if self.driver.wait_activity(".DialtactsActivity", 10):
            # try:

            #     keypad_tab = self.driver.find_element(
            #         AppiumBy.XPATH,
            #         "//android.widget.TextView[@resource-id='com.samsung.android.dialer:id/tab_text_view' and @text='Keypad']"
            #     )
            #     keypad_tab.click()
            #     print("Keypad tab clicked.")
            #     time.sleep(1)
            # except Exception as e:
            #     print(f"Failed to click Keypad tab: {e}")
            #     self.fail("Failed to click keypad tab.")

            # try:
            #     dialpad = self.driver.find_element(
            #         AppiumBy.ID, "com.samsung.android.dialer:id/dialpad_keypad_only"
            #     )
            #     print("Dialpad detected.")
            # except:
            #     print("Dialpad not found.")
            #     self.fail("Dialpad not visible.")
            
            expected_result = [
                "Welcome",
                "1: Login",
                "2: Exit"
            ]

            self.dialussdcode(expected_result)


            # pin = "*815#"
            # for digit in pin:  
            #     try:
            #         if digit == "*":
            #           key = self.driver.find_element(
            #           AppiumBy.XPATH,
            #           '//android.widget.RelativeLayout[@content-desc="Asterisk"]'
            #            )
            #         elif digit == "#":

                    
            #             key = self.driver.find_element(
            #             AppiumBy.XPATH,
            #             '//android.widget.RelativeLayout[@content-desc="Pound"]'
            #            )
            #         else:
            #             key = self.driver.find_element(
            #                 AppiumBy.XPATH,
            #                 f'//android.widget.TextView[@resource-id="com.samsung.android.dialer:id/dialpad_key_number" and @text="{digit}"]'
            #             )

            #         key.click()
            #         time.sleep(0.2)

            #     except Exception as e:
            #         print(f"Error pressing '{digit}': {e}")
            #         self.fail(f"Failed to enter digit '{digit}'")

            # try:
            #     phone_button = WebDriverWait(self.driver, 10).until(
            #     EC.element_to_be_clickable((AppiumBy.XPATH, '//android.widget.FrameLayout[@content-desc="Call button"]'))
            #     )
            #     print("Phone Button Detected")
            #     phone_button.click()

            # except:
            #     print("Waiting for USSD response...")
            #     time.sleep(5)

            try:

                ussd_screen = WebDriverWait(self.driver, 30).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.LinearLayout[@resource-id='android:id/parentPanel']"))
                )

                if ussd_screen:
                   print("BOA USSD Login Notification Detected", flush=True)
                else:
                    self.fail("No USSD screen detected")
                    self.driver.quit()

                input_field = self.driver.find_element(AppiumBy.XPATH, '//android.widget.EditText[@resource-id="com.android.phone:id/input_field"]')
                print("Input field detected", flush=True)
            except:
                print("USSD response not detected.", flush=True)
                self.fail("USSD screen did not appear.")

            try:
              
            #   option_number = input("Please Enter '1' to Login and '2' for Exit :").strip()
              option_number = '1'


              input_field.clear()

              if option_number == "2":
                  
                  input_field.send_keys(option_number)
                  send_button = WebDriverWait(self.driver, 30).until(
                  EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.Button[@resource-id='android:id/button1']")))
                  send_button.click()
                  print("Send Button Clicked Exiting...", flush=True)
                  time.sleep(0.3)

                  Snack_bar = WebDriverWait(self.driver, 30).until(
                  EC.element_to_be_clickable((AppiumBy.XPATH, '//android.widget.LinearLayout[@resource-id="android:id/parentPanel"]'))
                  )

                  if Snack_bar:
                    self.fail("Log out successfully")

                  else:
                      self.fail("Error while exiting try again pleace")
                      return

              else:

                  expected_Result = [
                        "Welcome to Bank of Abyssinia Mobile Banking Service.",
                        "1: Login",
                        "2: Exit"
                        ]

                  status = self.send_ussd(expected_Result ,'1', 'enter_pin_to_login')

                  if not status:
                       self.update_status("login", [2], "Fail", 7)
                       print("Welcome page not found Exiting..", flush=True)
                       self.cancel_ussd()
                       self.driver.quit()
                       exit()


                  print("Pin Enter Notification Detected", flush=True)
                    
                  attempt = 0
                  max_attempts = 2

                  while attempt < max_attempts:

                        if attempt == 1:               
                          print("Wrong pin", flush=True)

                        elif attempt == 2:
                            print("Too many failed attempts. User locked", flush=True)
                            self.driver.quit()
                            exit()
                            # print("Wrong one chance left")

                        # elif (attempt == 3):
                        #     print("Too many failed attempts. User locked")
                        #     self.driver.quit()
                        #     exit()

                        else:
                            pin_input = WebDriverWait(self.driver, 20).until(
                            EC.presence_of_element_located((AppiumBy.XPATH, '//android.widget.EditText[@resource-id="com.android.phone:id/input_field"]')))
                            # pin_number = input("Please enter your pin:")
                            pin_number = self.pin
                            
                            expected_Result = [
                                 "Please enter your PIN to login:"
                                ]

                            self.send_ussd(expected_Result ,pin_number,"enter_pin_to_login")

                            try:
                                message_element = WebDriverWait(self.driver, 20).until(
                                EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.TextView[@resource-id='com.android.phone:id/message']")))
                                message_text = message_element.text.strip()
                                    
                                if "Welcome" in message_text and "My Accounts" in message_text and "Transfer" in message_text:
                                    print("Login successful.", flush=True)
                                    break
                                else:
                                    print(f"Received response: {message_text}", flush=True)
                                    attempt += 1

                            except Exception as e:
                                print(f"Could not find message element: {e}", flush=True)
                                attempt += 1

                    # else:
                    #     self.fail("Pin enter notification not detected")
                        
                        return
            except:
                    self.fail("Pin enter notification not detected")
                    return
            
            ussd_message = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((AppiumBy.ID, 'com.android.phone:id/message'))).text
            
            expected_options = [
                        "1: My Accounts",
                        "2: Transfer",
                        "3: Transfer to Other Bank",
                        "4: Transfer to Own",
                        "5: Airtime",
                        "6: Utilities",
                        "7: Exchange Rates",
                        "8: More options"
                    ]
            
            not_found = []

            for option in expected_options:
                        if option in ussd_message:
                            print(f"Found option: {option}", flush=True)
                        else:
                            not_found.append(option)
                            print(f"Not Found option: {option}", flush=True)
                            self.fail(f"Missing option in USSD menu: {option}")

            self.update_status("login", [2], "Pass", 7)

            if(not_found):
              print(f"Not Found option: {not_found}", flush=True)

           
    def my_account(self):
            try:

                input_field = WebDriverWait(self.driver,20).until(
                    EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.EditText[@resource-id='com.android.phone:id/input_field']"))
                )
                input_field.clear()
                input_field.send_keys("1")

                send_key = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.Button[@resource-id='android:id/button1']")))
                send_key.click()

                message_element = WebDriverWait(self.driver, 20).until(
                    EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.TextView[@resource-id='com.android.phone:id/message']"))
                )

                message_text = message_element.text

                print("USSD Response:", message_text, flush=True)

                assert "My Accounts" in message_text , "Did not land on the expected 'My Accounts' page"

                accounts = [line for line in message_text.split("\n") if "ETB" in line or "Dollar" in line]
                print(accounts, flush=True)

                if not accounts:
                  self.update_status("Accounts", [2], "Fail", 7)
                  self.fail("No ETB or Dollar accounts found.")
                else:
                  print("list of accounts associated with the customer ID is displayed", flush=True)
                  self.update_status("Accounts", [2], "Pass", 7)
        
                  print("Excel status Updated", flush=True)
                print(f"{len(accounts)} account(s) found: {accounts}", flush=True)

                for i, acc in enumerate(accounts, start=1):
                    print(f"\nChecking Account {i}: {acc}", flush=True)

                    print(acc.split("-")[0].strip(), flush=True)

                    account_number_acc = acc.split(":")[1].split("-")[0].strip()

                    print(account_number_acc, flush=True)

                    self.send_ussd_input(str(i))

                    txn_page = self.get_ussd_message()

                    if txn_page :
                        assert account_number_acc in txn_page, "Account name mismatch on transaction page"

                        assert "Transactions" in txn_page, "Transaction list not found"

                        print("The list of recent transactions is displayed", flush=True)

                        self.update_status("Accounts", [4], "Pass", 7)

                    else:
                        print("Transaction page can't be verified or Transaction list not found", flush=True)
                        
                        self.update_status("Accounts", [4], "Fail", 7)

                    self.send_ussd_input("1")

                    txn_detail = self.get_ussd_message()

                    if txn_detail:
                        assert "debited from" in txn_detail or "credited to" in txn_detail, self.update_status("Accounts", [6], "Fail", 7); "No Transaction Details found"  # read and update the "status" column page 3 "Accounts" under the row A003 or row 5,6 and 7 to be "fail" and the cell 5,6 and 7 and from USSD_Test_Script.xlsx of the exel The list of recent transactions will be displayed
                        print("Transaction detail verified", flush=True)
                        # read and update the "status" column page 3 "Accounts" under the row A003 or row 5,6 and 7 to be "pass" and from USSD_Test_Script.xlsx of the exel The list of recent transactions will be displayed
                        self.update_status("Accounts", [6], "Pass", 7)
                    else:
                        # self.fail("Transaction detail page can't be verified")
                        print("Transaction detail page can't be verified", flush=True)
                        self.update_status("Accounts", [6], "Fail", 7)
                        # read and update the "status" column page 3 "Accounts" under the row A003 or row 5,6 and 7 to be "fail" and the cell 5,6 and 7 will be red and from USSD_Test_Script.xlsx of the exel The list of recent transactions will be displayed
                        

                    self.send_ussd_input("*")
                    back1 = self.get_ussd_message()
                    assert account_number_acc in back1, "Did not navigate back to transaction page1"
                    assert "Transactions" in back1, "Did not navigate back to transaction page2"

                    self.send_ussd_input("*")

                    back2 = self.get_ussd_message()

                    assert "My Accounts" in back2, "Did not navigate back to My Accounts page"

                    accounts = [line for line in message_text.split("\n") if "ETB" in line or "Dollar" in line]

                    if not accounts:
                      self.fail("No ETB or Dollar accounts found.")

                print("All account checks completed successfully", flush=True)

                self.send_ussd_input("*")

                back2 = self.get_ussd_message()

                ussd_message_home = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((AppiumBy.ID, 'com.android.phone:id/message'))).text
            
                expected_options = [
                        "1: My Accounts",
                        "2: Transfer",
                        "3: Transfer to Other Bank",
                        "4: Transfer to Own",
                        "5: Airtime",
                        "6: Utilities",
                        "7: Exchange Rates",
                        "8: More options"
                    ]
            
                not_found = []

                for option in expected_options:
                            if option in back2:
                                print(f"Found option: {option}", flush=True)
                            else:
                                not_found.append(option)
                                print(f"Not Found option: {option}", flush=True)
                                self.fail(f"Missing option in USSD menu: {option}")

                if(not_found):
                  print(f"Not Found option: {not_found}", flush=True)
                else:
                    print("Retuned to home page")

            except Exception as e:
                print("My accounts Test failed with error:", e, flush=True)

            print("My Account test compleated!", flush=True)
            self.cancel_ussd()
            time.sleep(1)

    def transfer(self):
        # transfer_withen_BOA = self.transfer_within_BOA()
        # if transfer_withen_BOA:
        #     print("Transfer with boa complated",flush=True)
        # else:
        #      print("Transfer with boa failed",flush=True)

        atm_withdrawal(self)

 
    def transfer_within_BOA(self):
        
        self.enter_pin_to_login()

        if not self.transfer_helper():
            print("Initial transfer helper failed.", flush=True)
            # return

        print("Test for Transfer within BoA", flush=True)


        expected_ResultB = [
            "Transfer",
            "1: Transfer within BoA", 
            "2: ATM withdrawal",
            "3: Load to TeleBirr",
            "4: Transfer to M-PESA",
            "5: Awach"
        ]
        
        WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
        )

        status = self.send_ussd(expected_ResultB, "1", "transfer_with_boa")

        if not status:
            self.update_status("Transfer", [2], "Fail", 7)
            # self.driver.quit()
            # return
            print("All expected value not found",flush=True)

        self.update_status("Transfer", [2], "Pass", 7)


        if not self.enter_account_and_select_valid_option(): #not handeld
            self.cancel_ussd()
            return
            

        print("Executing first round with user-input amount...", flush=True)

        if not self.handle_transfer_flow(is_negative_scenario=False, predefined_amount=self.amount, exceeded_amount=""):
            print("failed on first handel transfer call",flush=True)
            return
        
        self.send_ussd_input("*")
        time.sleep(1)

        fixed_amounts = [500000, 502000]  


        for idx, test_amt in enumerate(fixed_amounts, start=2): 
                print(f"Testing Negative Scenario Round {idx} with amount: {test_amt}", flush=True)

               
                if idx == 2:
                    if not self.transfer_helper():
                        print("Retrying transfer helper failed...", flush=True)
                        return

                    status = self.send_ussd(expected_ResultB, "1", "transfer_with_boa")
                    if not status:
                        return

                    if not self.enter_account_and_select_valid_option():
                        return

                result = self.handle_transfer_flow(
                    is_negative_scenario=True, 
                    predefined_amount="", 
                    exceeded_amount=test_amt, 
                    round_idx=idx  
                )

                if result == "retry":
                    print("Retrying the same round due to duplicate transaction error.", flush=True)
                    if not self.send_ussd_input("*"):
                        print("Failed to navigate back after duplicate transaction error", flush=True)
                    continue  

                if not result:
                    return 
        
        time.sleep(0.5)

        print("Transfer with BOA complated Successfully",flush=True)

        time.sleep(0.5)
        self.cancel_ussd()
        return True
    

    def transfer_helper(self):

        # self.enter_pin_to_login()

        print("Test for Transfer..", flush=True)

        expected_Result = [
              "1: My Accounts",
              "2: Transfer",
              "3: Transfer to Other Bank",
              "4: Transfer to Own",
              "5: Airtime",
              "6: Utilities",
              "7: Exchange Rates",
              "8: More options"
            ]
        
        WebDriverWait(self.driver, 50).until(
         EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message")))

        status = self.send_ussd(expected_Result ,"2","transfer")

        return status
    
    def airtime_helper(self):

        # self.enter_pin_to_login()

        print("Test for Airtime..", flush=True)

        expected_Result = [
              "1: My Accounts",
              "2: Transfer",
              "3: Transfer to Other Bank",
              "4: Transfer to Own",
              "5: Airtime",
              "6: Utilities",
              "7: Exchange Rates",
              "8: More options"
            ]
        
        WebDriverWait(self.driver, 50).until(
         EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message")))

        status = self.send_ussd(expected_Result ,"5","Airtime")

        return status
    
    def cancel_ussd(self):
                     
        try:
            cancel_button_locator = (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Cancel")')

            WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable(cancel_button_locator))
            cancel_button = self.driver.find_element(*cancel_button_locator)

            cancel_button.click()
                     
        except Exception as e:
            print(f"Failed to cancel': {e}", flush=True)
            self.driver.quit()
            exit()

    def handle_transfer_flow(self, is_negative_scenario, predefined_amount, exceeded_amount, round_idx=0):
            amount = predefined_amount if not is_negative_scenario else exceeded_amount

            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
            )

            print(f"Entering amount: {amount}", flush=True)
            if not self.send_ussd(["Enter Amount"], amount, "transfer_with_boa"):
                print("Amount entry failed", flush=True)
                return False

            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
            )
            print("Entering remark: testremark", flush=True)
            if not self.send_ussd(["Enter Remark"], "testremark", "transfer_with_boa"):
                print("Remark entry failed", flush=True)
                return False

            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
            )
            confirm_text = self.get_ussd_message()
            print(f"Confirmation screen:\n{confirm_text}", flush=True)

            if not self.send_ussd(["Please Confirm"], "1", "transfer_with_boa"):
                print("Confirmation failed", flush=True)
                return False

            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
            )
            final_text = self.get_ussd_message()
            print(f"Final response:\n{final_text}", flush=True)

            self.last_response = final_text  

           
            if "exceeded the daily transaction limit" in final_text.lower():
                print("Transfer failed due to daily limit  error.", flush=True)
                if is_negative_scenario:
                    self.update_status("Transfer", [8], "Pass", 7)
                    self.send_ussd_input("*")  
                    return True
                else:
                    self.update_status("Transfer", [8], "Fail", 7)
                    return False

            if "exceeded the single transaction limit" in final_text.lower():
                print("Transfer failed due to single limit.", flush=True)
                if is_negative_scenario:
                    self.update_status("Transfer", [9], "Pass", 7)
                    self.send_ussd_input("*")  
                    # self.cancel_ussd()
                    return True
                else:
                    self.update_status("Transfer", [9], "Fail", 7)
                    return False

            
            if "repository.productTransaction" in final_text or "productTransactionAlreadyExists" in final_text:
                print("Duplicate transaction or repo error. Will retry after going back.", flush=True)
                # self.send_ussd_input("*")
                return "retry"

            
            if "complete" in final_text.lower() or "debited" in final_text.lower():
                print("Transfer successfully", flush=True)
                self.update_status("Transfer", [5], "Pass", 7)
                return True

            print("Transfer response could not be validated.", flush=True)
            return False

    def enter_account_and_select_valid_option(self):
            expected_prompt = ["Enter Account No"]
            
            for attempt in range(1, 4):
                time.sleep(0.5)
                accountno = self.account_number 
                print(f"Attempt {attempt}: Entering account number {accountno}", flush=True)

                WebDriverWait(self.driver, 20).until(
                    EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
                )
                
                status = self.send_ussd(expected_prompt, accountno, "transfer_with_boa")

                if not status:
                    print("Account number entry failed.", flush=True)
                    return False
                
               
                txn_text = self.get_ussd_message()
                print(f"USSD message after entering account:\n{txn_text}", flush=True)


                account_option, detected_account = self.extract_staff_or_saving_option(txn_text)

                if not account_option or not detected_account:
                    print("No valid STAFF or SAVING account detected. Retrying...", flush=True)
                    self.send_ussd_input("*")  
                    continue
                
                #not handeld
                if accountno == detected_account: # not handeld
                    print("Error: Source and destination accounts are the same. Retrying...", flush=True)
                    self.send_ussd_input("*")
                    if attempt == 3:
                        self.cancel_ussd()
                        return False
                    continue

                WebDriverWait(self.driver, 20).until(
                    EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
                )
                
                self.send_ussd(["Transfer within BoA"], account_option, "transfer_with_boa")
                return True

            print("Failed to enter a valid account after 3 attempts.", flush=True)
            return False

    def extract_staff_or_saving_option(self, ussd_text):
        lines = ussd_text.split('\n')
        for line in lines:
            if "STAFF" in line.upper() or "SAVING" in line.upper():
                if ":" in line:
                    option_number =  line.split(":")[0].strip()
                    account_number_ext = line.split(":")[1].split("-")[0].strip()
                    return option_number, account_number_ext
        return None, None
    
    def send_ussd(self, expected_values, value, function):
      try:
          if self.wait_for_expected_output(expected_values):
            
                    input_field = WebDriverWait(self.driver, 25).until(
                        EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.EditText[@resource-id='com.android.phone:id/input_field']"))
                    )

                    input_field.clear()
                    input_field.send_keys(value)

                    send_button_locator = (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Send")')

                    WebDriverWait(self.driver, 25).until(EC.element_to_be_clickable(send_button_locator))
                    send_button = self.driver.find_element(*send_button_locator)

                    send_button.click()
                    return True

          else:
            print("Mismatch in expected output.", flush=True)

            if function == "enter_pin_to_login":
                print("Exiting due to failure in login step.", flush=True)
                self.driver.quit()
                exit()
            else:
                return False

      except Exception as e:
        print(f"Error during USSD input: {e}", flush=True)
        if function == "enter_pin_to_login":
            self.driver.quit()
            exit()
        else:
            return False

    def wait_for_expected_output(self, expected_values, timeout=25):
  
            try:
                WebDriverWait(self.driver, timeout).until(
                    EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
                )
                message_element = self.driver.find_element(AppiumBy.ID, "com.android.phone:id/message")
                ussd_text = message_element.text.strip().lower()
                # print("USSD Output:\n", ussd_text)

                for expected in expected_values:
                    if expected.lower() not in ussd_text:
                        print(f"Expected text not found: '{expected}'", flush=True)
                        return False

                return True
            
            except Exception as e:
                print(f"Error waiting for USSD message: {e}", flush=True)
                return False

    def extract_and_validate_success_screen(self, message):

        try:
            message = message.lower()

            # amount_match = re.search(r'etb\s+([\d,.]+)', message)
            amount_match = re.search(r'etb\s*([\d,.]+)', message)
            from_acc_match = re.search(r'debited\s+from\s+(\d+)', message)
            to_acc_match = re.search(r'for\s+(\d+)', message)
            txn_id_match = re.search(r'id:\s*([a-z0-9]+)', message)

            if amount_match and from_acc_match and to_acc_match:
                print(f"Amount: {amount_match.group(1)} ETB", flush=True)
                print(f"From Account: {from_acc_match.group(1)}", flush=True)
                print(f"To Account: {to_acc_match.group(1)}", flush=True)
                if txn_id_match:
                    print(f"Transaction ID: {txn_id_match.group(1).upper()}", flush=True)
                else:
                    print("Transaction ID not found in message.", flush=True)

                return True
            else:
                print("Could not extract all fields from the message.", flush=True)
                return False
        except Exception as e:
            print(f"Error while parsing success screen: {e}", flush=True)
            return False

    def send_ussd_input(self, value):

        input_field = WebDriverWait(self.driver, 25).until(
            EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.EditText[@resource-id='com.android.phone:id/input_field']"))
        )

        input_field.clear()
        input_field.send_keys(value)

        # send_key = WebDriverWait(self.driver, 10).until(
        #     EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.Button[@resource-id='android:id/button1']"))
        # )

        send_button_locator = (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Send")')

        WebDriverWait(self.driver, 25).until(EC.element_to_be_clickable(send_button_locator))
        send_button = self.driver.find_element(*send_button_locator)

        send_button.click()

        # send_key.click()

    def get_ussd_message(self):
        
        message_element = WebDriverWait(self.driver, 25).until(
            EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.TextView[@resource-id='com.android.phone:id/message']"))
        )
        msg = message_element.text.strip()
        print("USSD Message:", msg, flush=True)
        
        return msg

    def get_resource_path(self, relative_path):
        try:
            # If bundled by PyInstaller
            base_path = sys._MEIPASS
        except AttributeError:
            # If running as a normal .py file
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)
    
    def update_status(self, sheet_name, row_indices, status, status_col_index):
        base_dir = os.path.dirname(os.path.abspath(__file__))  
        file_path = "USSD_Test_Script.xlsx"  
        # file_path = os.path.join(base_dir, 'USSD_Test_Script.xlsx')
        # file_path = self.get_resource_path("USSD_Test_Script.xlsx")
        
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        no_fill = PatternFill(fill_type=None)

        for row in row_indices:
            cell = ws.cell(row=row, column=status_col_index)

            if isinstance(cell, MergedCell):
                print(f"Skipping merged cell at row {row}, column {status_col_index}", flush=True)
                continue

            print(f"Writing '{status}' to row {row}, col {status_col_index}", flush=True)
            cell.value = status

            if status.lower() == "fail":
                cell.fill = red_fill
            else:
                cell.fill = no_fill

        wb.save(file_path)
        print(f"Updated '{sheet_name}' sheet - rows {row_indices} set to '{status}'.", flush=True)
            
                    
    def test_app(self):
        try:
            # self.enter_pin_to_login()
            # my_account = self.my_account()
            # transfer = self.transfer()
            # airtime_topup_for_ethio_telecom(self)
            # airtime_topup_for_safaricom(self)
            transfer_to_own_account(self)
        except Exception as e:
            print(f"Error: {e}", flush=True)
            self.fail("USSD process failed.")


if __name__ == '__main__':
    # unittest.main()
    unittest.main(argv=['first-arg-is-ignored'], exit=False)

# appium --allow-insecure adb_shell

# pyinstaller --onefile --noconsole  --add-data "test_app2.py;." tkintergui.py
