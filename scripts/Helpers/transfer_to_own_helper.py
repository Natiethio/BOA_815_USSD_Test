import unittest
import time
import pandas as pd
from bs4 import BeautifulSoup
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from appium.options.android import UiAutomator2Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import re
import sys
import os


def transfer_to_own_helper(self):

        # self.enter_pin_to_login()

        print("Test for Transfer to Own..", flush=True)

        expected_Result = [
              "1: My Accounts",
              "2: Transfer",
              "3: Transfer to Other Bank",
              "4: Transfer to Own",
              "5: Airtime",
              "6: Utilities",
              "7: Exchange Rates",
              "8: More options"
            ]
        
        WebDriverWait(self.driver, 50).until(
         EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message")))

        status = self.send_ussd(expected_Result ,"4","Transfer to Own")

        return status