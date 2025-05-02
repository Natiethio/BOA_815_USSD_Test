import os
import sys
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def get_executable_dir():
    """Returns the folder where the .exe or .py is running."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable) 
    else:
        return os.path.dirname(os.path.abspath(__file__))

def ensure_excel_file_exists():
    """Copies the bundled Excel file to the executable directory if it's not already there."""
    exe_dir = get_executable_dir()
    target_path = os.path.join(exe_dir, "Test.xlsx")

    if not os.path.exists(target_path):
        try:
            
            bundled_path = os.path.join(sys._MEIPASS, "Test.xlsx")
            shutil.copy(bundled_path, target_path)
            print(f"Excel file copied to: {target_path}")
        except Exception as e:
            print(f"Failed to copy Excel file: {e}")

    return target_path

def write_hello_to_excel():
    excel_path = ensure_excel_file_exists()

    try:
        wb = load_workbook(excel_path)
        ws = wb["Second"]
        ws.cell(row=3, column=5).value = "Hello"
        wb.save(excel_path)
        print(f"'Hello' written to {excel_path}, Sheet='Second', Row=3, Col=5")
        print(f"Saving Excel file to: {excel_path}")
    except Exception as e:
        print(f"Error writing to Excel: {e}")

if __name__ == "__main__":
    write_hello_to_excel()