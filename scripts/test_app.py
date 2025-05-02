import unittest
import time
import pandas as pd
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from appium.options.android import UiAutomator2Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import re
import sys
import os
from Transfer.transfer import transfer
import shutil
import Helpers.android_helper as android_helper
from datetime import datetime
from AirTime.airtime import airtime
import subprocess
from Transfer_otherBank.otherbank import otherbank
from My_Account.my_account import my_account
from Transfer_to_own.transfer_to_own import transfer_to_own_account



class TestAppium(unittest.TestCase):

    ussd_log_path = None
    ussd_log_initialized = False
    slow_popups = []
    all_slow_popups = []

    auto_device_id = android_helper.get_connected_device()
    auto_android_ver = android_helper.get_android_version(auto_device_id) if auto_device_id else None
    pkg, main_activity = android_helper.get_dialer_info()
    # device_id      = sys.argv[1] if len(sys.argv) > 1 else "R9ZR601C18H"
    # android_ver    = sys.argv[2] if len(sys.argv) > 2 else "12"
    pin            = sys.argv[1] if len(sys.argv) > 1 else ""
    name           = sys.argv[2] if len(sys.argv) > 2 else "Sender Name"
    account_number = sys.argv[3] if len(sys.argv) > 3 else "174615624"
    amount         = sys.argv[4] if len(sys.argv) > 4 else "5"
    etl_number   = sys.argv[5] if len(sys.argv) > 5 else "0970951608"
    safaricom_number   = sys.argv[6] if len(sys.argv) > 6 else "0712911008"
    bank_account = sys.argv[7] if len(sys.argv) > 7 else " "
    bank_name = sys.argv[8] if len(sys.argv) > 8 else "CBE"
    module_name = sys.argv[9] if len(sys.argv) > 9 else "All"

    def get_connected_device(self):
        try:
            result = subprocess.run(['adb', 'devices'], capture_output=True, text=True, check=True)
            output_lines = result.stdout.strip().split('\n')
            if len(output_lines) > 1:
                device_line = output_lines[1].split('\t')
                if len(device_line) == 2 and device_line[1] == 'device':
                    return device_line[0]
            return None
        except subprocess.CalledProcessError as e:
            print(f"Error running ADB devices: {e}")
            self.driver.quit()
            exit()
            return None
             
    # auto_device_id = get_connected_device()
    auto_android_ver = android_helper.get_android_version(auto_device_id) if auto_device_id else None

    print(f"Using Device ID: {auto_device_id if auto_device_id else "qewrtwtey"}")
    print(f"Using Android Version: {auto_android_ver if auto_android_ver else "12"}")
    print(f"Using package: {pkg}")
    print(f"Using Main_Activity: {main_activity}")

    def setUp(self) -> None:
        options = UiAutomator2Options()
        options.platform_name = "Android"
        options.automation_name = "UiAutomator2"
        options.device_name = self.auto_device_id 
        options.udid = self.auto_device_id 
        options.platform_version = self.auto_android_ver
        options.app_package = self.pkg
        options.app_activity = self.main_activity
        options.no_reset = True
        options.language = "en"
        options.locale = "US"
        self.init_ussd_log_file()
        self.excel_path=None
        self.reports_dir=None

        try:
            self.driver = webdriver.Remote("http://127.0.0.1:4723", options=options)
            self.driver.implicitly_wait(5)
        except Exception as e:
           print("Error initializing appium exiting...",e)
           self.driver.quit()
           exit()

    def tearDown(self) -> None:
        if self.driver:
            self.driver.quit()

    def init_ussd_log_file(self):
        log_dir = os.path.join(os.getcwd(), "popuptime")
        os.makedirs(log_dir, exist_ok=True)
        # os.makedirs("popuptime", exist_ok=True)

        if not self.ussd_log_initialized:
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            self.ussd_log_path = f"popuptime/popup_page_{timestamp}.txt"
            self.ussd_log_initialized = True

        with open(self.ussd_log_path, "a", encoding="utf-8") as f:
            f.write("\n========================== USSD Pop Up pages time Logs ========================\n\n")

    def dialussdcode(self,expected_texts, retries=5, wait_time=7):
       
        ussd_code = '*815' + '%23'

        for attempt in range(1, retries + 1):
            print(f"Dialing USSD code (Attempt {attempt})...", flush=True)
            self.driver.execute_script('mobile: shell', {
                'command': 'am',
                'args': ['start', '-a', 'android.intent.action.CALL', 'tel:' + ussd_code]
            })

            try:
                message_element = WebDriverWait(self.driver, wait_time).until(
                    EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
                )
                ussd_text = message_element.text.strip()
                # print("USSD Response:", ussd_text, flush=True)

                if all(expected.lower() in ussd_text.lower() for expected in expected_texts):
                    print("Expected USSD page matched.", flush=True)
                    return True
                else:
                    self.cancel_ussd()
                    time.sleep(0.3)

                print("Expected USSD response not found. Retrying...", flush=True)

            except Exception as e:
                print(f"Error waiting for USSD response: {e}", flush=True)

            time.sleep(2)  

        self.fail(" Failed to load the expected USSD screen after retries.")


    def enter_pin_to_login(self):

            
            expected_result = [
                "Welcome",
                "1: Login",
                "2: Exit"
            ]

            self.dialussdcode(expected_result)


 
            try:

                # ussd_screen = WebDriverWait(self.driver, 20).until(
                # EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.LinearLayout[@resource-id='android:id/parentPanel']"))
                # )

                # if ussd_screen:
                #    print("BOA USSD Login Notification Detected", flush=True)
                # else:
                #     self.fail("No USSD screen detected")
                #     self.driver.quit()

                input_field = self.driver.find_element(AppiumBy.XPATH, '//android.widget.EditText[@resource-id="com.android.phone:id/input_field"]')
                print("Input field detected", flush=True)
            except:
                print("USSD response not detected.", flush=True)
                self.fail("USSD screen did not appear.")

            try:
              
            #   option_number = input("Please Enter '1' to Login and '2' for Exit :").strip()
              option_number = '1'


              input_field.clear()

              if option_number == "2":
                  
                  input_field.send_keys(option_number)
                  send_button = WebDriverWait(self.driver, 30).until(
                  EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.Button[@resource-id='android:id/button1']")))
                  send_button.click()
                  print("Send Button Clicked Exiting...", flush=True)
                  time.sleep(0.3)

                  Snack_bar = WebDriverWait(self.driver, 30).until(
                  EC.element_to_be_clickable((AppiumBy.XPATH, '//android.widget.LinearLayout[@resource-id="android:id/parentPanel"]'))
                  )

                  if Snack_bar:
                    self.fail("Log out successfully")

                  else:
                      self.fail("Error while exiting try again pleace")
                      return

              else:

                  expected_Result = [
                        "Welcome to Bank of Abyssinia Mobile Banking Service.",
                        "1: Login",
                        "2: Exit"
                        ]

                  status = self.send_ussd(expected_Result ,'1', 'enter_pin_to_login', "Enter Pin Page")

                  if not status:
                       self.update_status("login", [2], "Fail", 5 ,screenshot_name="enter_pin_no_welcomepage")
                       print("Welcome page not found Exiting..", flush=True)
                       self.cancel_ussd()
                       self.driver.quit()
                       exit()


                  print("Pin Enter Notification Detected", flush=True)
                    
                  attempt = 0
                  max_attempts = 3

                  while attempt < max_attempts:

                        if attempt == 1:               
                          print("Wrong pin", flush=True)

                        elif attempt == 2:
                            print("Too many failed attempts with wrong. Canceld", flush=True)
                            self.cancel_ussd()
                            self.driver.quit()
                            exit()
                            # print("Wrong one chance left")

                        # elif (attempt == 3):
                        #     print("Too many failed attempts. User locked")
                        #     self.driver.quit()
                        #     exit()

                        else:
                            pin_input = WebDriverWait(self.driver, 20).until(
                            EC.presence_of_element_located((AppiumBy.XPATH, '//android.widget.EditText[@resource-id="com.android.phone:id/input_field"]')))
                            # pin_number = input("Please enter your pin:")
                            pin_number = self.pin
                            
                            expected_Result = [
                                 "Please enter your PIN to login:"
                                ]

                            self.send_ussd(expected_Result ,pin_number,"enter_pin_to_login", "Home Page")

                            try:
                                message_element = WebDriverWait(self.driver, 20).until(
                                EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.TextView[@resource-id='com.android.phone:id/message']")))
                                message_text = message_element.text.strip()
                                    
                                if "Welcome" in message_text and "My Accounts" in message_text and "Transfer" in message_text:
                                    print("Login successful.", flush=True)
                                    break
                                else:
                                    print(f"Received response: {message_text}", flush=True)
                                    attempt += 1

                            except Exception as e:
                                print(f"Could not find message element: {e}", flush=True)
                                attempt += 1

                    # else:
                    #     self.fail("Pin enter notification not detected")
                        
                        return
            except:
                    self.fail("Pin enter notification not detected")
                    return
            
            ussd_message = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((AppiumBy.ID, 'com.android.phone:id/message'))).text
            
            expected_options = [
                        "1: My Accounts",
                        "2: Transfer",
                        "3: Transfer to Other Bank",
                        "4: Transfer to Own",
                        "5: Airtime",
                        "6: Utilities",
                        "7: Exchange Rates",
                        "8: More options"
                    ]
            
            not_found = []

            for option in expected_options:
                        if option in ussd_message:
                            print(f"Found option: {option}", flush=True)
                        else:
                            not_found.append(option)
                            print(f"Not Found option: {option}", flush=True)
                            self.fail(f"Missing option in USSD menu: {option}")

            self.update_status("login", [2], "Pass", 5) #pass

            if(not_found):
              print(f"Not Found option: {not_found}", flush=True)

    def cancel_ussd(self):
                     
        try:
            cancel_button_locator = (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Cancel")')

            WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable(cancel_button_locator))
            cancel_button = self.driver.find_element(*cancel_button_locator)

            cancel_button.click()
            print("USSD session cancelled successfully.", flush=True)

        except Exception as e:
            print(f"Failed to cancel': {e}", flush=True)
            self.driver.quit()
            exit()
 
    def send_ussd(self, expected_values, value, function, next_page=None):
      time_out = 25
      try:
          if self.wait_for_expected_output(expected_values):
            
                    input_field = WebDriverWait(self.driver, 25).until(
                        EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.EditText[@resource-id='com.android.phone:id/input_field']"))
                    )

                    input_field.clear()
                    input_field.send_keys(value)

                    send_button_locator = (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Send")')

                    WebDriverWait(self.driver, 25).until(EC.element_to_be_clickable(send_button_locator))
                    send_button = self.driver.find_element(*send_button_locator)

                    if next_page:
                      start_time = time.time()

                    send_button.click()

                    if WebDriverWait(self.driver, time_out).until(
                    EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))):
                        
                            if next_page:
                                end_time = time.time()
                                duration = round(end_time - start_time, 2)

                                with open(self.ussd_log_path, "a", encoding="utf-8") as f:
                                    f.write(f"{next_page} page appeared in => {duration} seconds\n")

                                if duration > 2.5:
                                   self.slow_popups.append((next_page, duration))
                                   self.all_slow_popups.append((next_page, duration))
                            return True
                    else:
                        print("Expected page doesn't apper in 25 second range", flush=True)
                        return False

          else:
            print("Mismatch in expected output.", flush=True)

            if function == "enter_pin_to_login":
                print("Exiting due to failure in login step.", flush=True)
                self.driver.quit()
                exit()
            else:
                return False

      except Exception as e:
        print(f"Error during USSD input: {e}", flush=True)
        if function == "enter_pin_to_login":
            self.driver.quit()
            exit()
        else:
            return False

    def wait_for_expected_output(self, expected_values, timeout=25):
  
            try:
                WebDriverWait(self.driver, timeout).until(
                    EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message"))
                )
                message_element = self.driver.find_element(AppiumBy.ID, "com.android.phone:id/message")
                ussd_text = message_element.text.strip().lower()
                # print("USSD Output:\n", ussd_text)

                for expected in expected_values:
                    if expected.lower() not in ussd_text:
                        print(f"Expected text not found: '{expected}'", flush=True)
                        return False

                return True
            
            except Exception as e:
                print(f"Error waiting for USSD message: {e}", flush=True)
                return False
            
    def log_popup_time(self, popup_text, time_taken):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_dir = os.path.join(os.getcwd(), "popuptime", f"popup_page_{timestamp[:8]}")
            os.makedirs(log_dir, exist_ok=True)

            log_file = os.path.join(log_dir, f"popup_page_{timestamp}.txt")
            with open(log_file, "a", encoding="utf-8") as f:
                f.write(f'"{popup_text}" appeared in => {time_taken} seconds\n')
            print(f"Popup timing logged: {popup_text} => {time_taken}s", flush=True)

        except Exception as e:
            print(f"Error logging popup time: {e}", flush=True)

    def send_ussd_input(self, value):

        input_field = WebDriverWait(self.driver, 25).until(
            EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.EditText[@resource-id='com.android.phone:id/input_field']"))
        )

        input_field.clear()
        input_field.send_keys(value)

        # send_key = WebDriverWait(self.driver, 10).until(
        #     EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.Button[@resource-id='android:id/button1']"))
        # )

        send_button_locator = (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Send")')

        WebDriverWait(self.driver, 25).until(EC.element_to_be_clickable(send_button_locator))
        send_button = self.driver.find_element(*send_button_locator)

        send_button.click()

        # send_key.click()

    def get_ussd_message(self):
        
        message_element = WebDriverWait(self.driver, 25).until(
            EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.TextView[@resource-id='com.android.phone:id/message']"))
        )
        msg = message_element.text.strip()
        # print("USSD Message:", msg, flush=True)
        
        return msg
 
    def get_executable_dir(self):
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        else:
            return os.path.dirname(os.path.abspath(__file__))
    
    def ensure_excel_file_exists(self):
            

            if self.excel_path:  # if already set, just reuse it
               return self.excel_path
            
            # exe_dir = self.get_executable_dir()
            exe_dir = os.getcwd()
            # report_dir = os.path.join(os.getcwd(), "reports")
            self.reports_dir = os.path.join(exe_dir, "reports")
            os.makedirs(self.reports_dir, exist_ok=True) 

            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            self.excel_path = os.path.join(self.reports_dir, f"USSD_Test_Script_{timestamp}.xlsx")
            # target_path = os.path.join(exe_dir, "USSD_Test_Script.xlsx")

            if not os.path.exists(self.excel_path ):
                try:
                    if hasattr(sys, '_MEIPASS'):
                # Running from a bundled executable
                        bundled_excel = os.path.join(sys._MEIPASS, "USSD_Test_Script2.xlsx")
                        shutil.copy(bundled_excel, self.excel_path)
                        print(f"Excel file copied to: {self.excel_path}", flush=True)
                    else:
                            # Running from a .py file
                            # print("from else")
                            source_excel = os.path.join(exe_dir, "USSD_Test_Script2.xlsx")
                            if os.path.exists(source_excel):
                                # shutil.copy(source_excel, self.excel_path)
                                self.excel_path = source_excel
                                print(f"Excel in the path is used", flush=True)
                            else:
                                raise FileNotFoundError(f"Source Excel file not found at {source_excel}")

                except Exception as e:
                        print(f"Failed to copy Excel file: {e}", flush=True)

            return self.excel_path
    
            
    def update_status(self, sheet_name, row_indices, status, status_col_index,screenshot_name=None):
        excel_paths = self.ensure_excel_file_exists()
        
        try:
            wb = load_workbook(excel_paths)
            ws = wb[sheet_name]

            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            no_fill = PatternFill(fill_type=None)

            for row in row_indices:

                target_row = row
                target_col = status_col_index

                # cell = ws.cell(row=row, column=status_col_index)
                cell = ws.cell(row=target_row, column=target_col)

                # if isinstance(cell, MergedCell):
                #     print(f"Skipping merged cell at row {row}, column {status_col_index}", flush=True)
                #     continue

                # print(f"Writing '{status}' to row {row}, col {status_col_index}", flush=True)
                # cell.value = status

                if isinstance(cell, MergedCell):
                    found_range = False
                    for merged_range in ws.merged_cells.ranges:
                        if (row, status_col_index) in merged_range:
                            top_left = (merged_range.min_row, merged_range.min_col)
                            cell = ws.cell(*top_left)
                            print(f"Cell at row {row}, col {status_col_index} is merged. Writing to top-left cell at {top_left}.", flush=True)
                            found_range = True
                            break

                    if not found_range:
                        print(f"Skipping merged cell at row {row}, column {status_col_index} (not found in any range)", flush=True)
                        continue

                else:
                    print(f"Writing '{status}' to row {row}, col {status_col_index}", flush=True)

                cell.value = status

                if status.lower() == "fail" and screenshot_name:

                    cell.fill = red_fill
                    failed_dir = os.path.join(os.getcwd(), "Failed")
                    os.makedirs(failed_dir, exist_ok=True)

                    timestamp = time.strftime("%Y%m%d-%H%M%S") 
                    filename = f"{screenshot_name}_{timestamp}.jpg"
                    filepath = os.path.join(failed_dir, filename)

                    self.driver.save_screenshot(filepath)
                    print(f"Screenshot saved: {filepath}", flush=True)

                else:
                    cell.fill = no_fill

            wb.save(excel_paths)
            print(f"Saving Excel file to: {excel_paths}", flush=True)
            print(f"Updated '{sheet_name}' sheet - rows {row_indices} set to '{status}'.", flush=True)
        
        except Exception as e:
            print(f"Error updating Excel file: {e}", flush=True)  

        # Any other initialization can go here
    def test_app(self):
        try:
            # self.enter_pin_to_login()
            if self.module_name == "1":
                my_account_opt = my_account(self)  
            elif self.module_name == "2":  
                transfer_one = transfer(self)  
            elif self.module_name == "3":  
                transfer_otherbank = otherbank(self)  
            elif self.module_name == "4":  
                transfer_toown = transfer_to_own_account(self)  
            elif self.module_name == "5":  
                airtimeres = airtime(self)  
            # elif self.module_name == "6":  
            #     utilities = utilities(self)  
            else:
                print("End to end test")  
                self.enter_pin_to_login()
                # my_account_opt = my_account(self)
                # # transfer_one = transfer(self)
                # # transfer_otherbank = otherbank(self)
                # # transfer_toown = transfer_to_own_account(self)
                airtimeres = airtime(self)

            if self.all_slow_popups:
                print("\n========= Cumulative Delayed USSD Pages ( > 2.5s ) =========")
                for page, duration in self.all_slow_popups:
                    print(f"{page}: {duration}s")

        except Exception as e:
            print(f"Error: {e}", flush=True)
            self.fail("USSD process failed.")


if __name__ == '__main__':
    # unittest.main()
    unittest.main(argv=['first-arg-is-ignored'], exit=False)

# appium --allow-insecure adb_shell

# pyinstaller --onefile --noconsole  --add-data "test_app2.py;." tkintergui.py
#  264 301