import unittest
import time
import pandas as pd
from bs4 import BeautifulSoup
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from appium.options.android import UiAutomator2Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import re
import sys
import os
from scripts.test_app import TestAppium



def atm_withdrawal(self):

        self.enter_pin_to_login()

        status_helper = self.transfer_helper()

        time.sleep(5)
        print("Test for Atm With drawal ..", flush=True)

        status_helper = self.transfer_helper()

        if not status_helper:
            print("No home page found retrying")
            self.cancel_ussd()
            self.atm_withdrawal()
            return
                
        expected_ResultB = [
            "Transfer",
            "1: Transfer within BoA", 
            "2: ATM withdrawal",
            "3: Load to TeleBirr",
            "4: Transfer to M-PESA",
            "5: Awach"
            ]
        
        expected_ResultC = [ 
            "ATM withdrawal"
        ]
                
        # WebDriverWait(self.driver, 20).until(
        # EC.presence_of_element_located((AppiumBy.ID, "com.android.phone:id/message")))

        status = self.send_ussd(expected_ResultB ,"2","atm_withdrawal")

        message_text = self.get_ussd_message()

        # assert "ATM withdrawal" in message_text , "Did not land on the expected 'ATM withdrawal' page"

        accounts = [line for line in message_text.split("\n") if "ETB" in line or "Dollar" in line]

        if not accounts:
            self.fail("No ETB or Dollar accounts found.")
        else:
            print(f"{len(accounts)} account(s) found: {accounts}", flush=True)

        for index, account in enumerate(accounts):
            print(f"\nChecking account: {account.strip()}", flush=True)
            selection = str(index + 1)

            status = self.send_ussd("ATM withdrawal", selection, "atm_withdrawal")

            if not status:
                print(f"Failed to select account at index {selection}", flush=True)
                continue

            message_text = self.get_ussd_message()

            account_number = account.split(":")[1].strip().split("-")[0].strip()

            status = self.send_ussd("Enter Amount(multiple of 100)", "100", "atm_withdrawal")

            if not status:
              print("Failed to send amount for ATM withdrawal", flush=True)
              continue


            message_text = self.get_ussd_message()
            # status = self.send_ussd("Please Confirm", "1", "atm_withdrawal")
            status = self.send_ussd("Please Confirm", "*", "atm_withdrawal")


            if not status:
               print("Transaction confirmation failed", flush=True)
               self.update_status("Transfer", [12], "Fail", 7)
               continue
            else:
                self.update_status("Transfer", [12], "Pass", 7)
                time.sleep(0.5)
                status = self.send_ussd("Enter Amount(multiple of 100)", "*", "atm_withdrawal")
            
            # status = self.send_ussd("Complete", "*", "atm_withdrawal")
            # print("Atm Transfer Successfully for", flash=True)



            # if status:
            #     self.update_status("Transfer", [12], "Pass", 7)
            # else:
            #     self.update_status("Transfer", [12], "Fail", 7)

            if index < len(accounts) - 1:
                    # status = self.send_ussd(expected_ResultB, "2", "atm_withdrawal")
                    status = self.send_ussd(expected_ResultB, "2", "atm_withdrawal")
                    if not status:
                        print("Failed to return to ATM withdrawal menu for next account", flush=True)
                        break
       
        self.cancel_ussd()